from django.shortcuts import redirect, render, get_object_or_404
from openpyxl import load_workbook
from organization.models import Branch, EndDayRecord
from user.permission import IsAdminAccountingOrStoreKeeperMixin

from django.urls import reverse_lazy
from django.db.utils import IntegrityError
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from user.permission import IsAdminMixin
from .models import ProductCategory, ItemReconcilationApiItem
from bill.models import Bill
from .forms import ProductCategoryForm
from django.contrib import messages
from datetime import date, datetime

class ProductCategoryMixin():
    model = ProductCategory
    form_class = ProductCategoryForm
    paginate_by = 50
    queryset = ProductCategory.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("product_category_list")
    search_lookup_fields = [
        "title",
        "description",
    ]


class ProductCategoryList(ProductCategoryMixin, ListView):
    template_name = "productcategory/productcategory_list.html"
    queryset = ProductCategory.objects.filter(status=True, is_deleted=False)


class ProductCategoryDetail(ProductCategoryMixin, DetailView):
    template_name = "productcategory/productcategory_detail.html"


class ProductCategoryCreate(ProductCategoryMixin, CreateView):
    template_name = "create.html"


class ProductCategoryUpdate(ProductCategoryMixin, UpdateView):
    template_name = "update.html"


class ProductCategoryDelete(ProductCategoryMixin, DeleteMixin, View):
    pass


from django.db import transaction
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import Product
from .forms import ProductForm


class ProductMixin():
    model = Product
    form_class = ProductForm
    paginate_by = 50
    queryset = Product.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("product_list")
    search_lookup_fields = [
        "title",
        "description",
    ]


class ProductList(ProductMixin, ListView):
    template_name = "product/product_list.html"
    queryset = Product.objects.filter(status=True, is_deleted=False)


class ProductDetail(ProductMixin, DetailView):
    template_name = "product/product_detail.html"


class ProductCreate(ProductMixin, CreateView):
    template_name = "product/product_create.html"


class ProductUpdate(ProductMixin, UpdateView):
    template_name = "update.html"


class ProductDelete(ProductMixin, DeleteMixin, View):
    pass

class ProductUploadView(View):

    def post(self, request):
        file = request.FILES.get('file', None)
        if not file:
            messages.error(request, 'Please Provide the correct file ')
            return redirect(reverse_lazy("product_create"))
        
        file_ext = file.name.split('.')[-1]
        if file_ext not in ['xlsx', 'xls']:
            messages.error(request, 'Format must be in xlsx or xls ')
            return redirect(reverse_lazy("product_create"))

        wb = load_workbook(file)
        try:
            food_category = ProductCategory.objects.get(title='FOOD')
            beverage_category= ProductCategory.objects.get(title='BEVERAGE')
            others_category = ProductCategory.objects.get(title='OTHERS')
        except ProductCategory.DoesNotExist:
            messages.error(request, 'Please Create Product Categories first')
            return redirect(reverse_lazy("product_create"))
        excel_data = list()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(cell.value)
                row_data.insert(0, sheet.title)
                excel_data.append(row_data)
        
        for data in excel_data:
            if not all(data[0:9]):
                continue
            if data[1].lower() == 'group':
                continue
            try:
                product = Product.objects.get(title__iexact=data[2].strip())
                product.group = data[1].strip()
                product.price = data[3]
                product.unit = data[4].strip()
                product.is_taxable = True if data[5].lower().strip() == "yes" else False
                product.is_produced = True if data[6].lower().strip() == "yes" else False
                product.reconcile = True if data[7].lower().strip() == "yes" else False
                product.is_billing_item = True if data[8].lower().strip() == "yes" else False

                if "food" in data[0].lower().strip():
                    product.type = food_category
                elif "beverage" in data[0].lower().strip():
                    product.type = beverage_category
                else:
                    product.type = others_category
                product.save()
            except Product.DoesNotExist:
                product = Product()
                product.group = data[1].strip()
                product.title=data[2].strip()
                product.price = data[3]
                product.unit = data[4].strip()
                product.is_taxable = True if data[5].lower().strip() == "yes" else False
                product.is_produced = True if data[6].lower().strip() == "yes" else False
                product.reconcile = True if data[7].lower().strip() == "yes" else False
                product.is_billing_item = True if data[8].lower().strip() == "yes" else False

                if "food" in data[0].lower().strip():
                    product.type = food_category
                elif "beverage" in data[0].lower().strip():
                    product.type = beverage_category
                else:
                    product.type = others_category
                product.save()
        return redirect(reverse_lazy("product_list"))

from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import CustomerProduct
from .forms import CustomerProductForm


class CustomerProductMixin():
    model = CustomerProduct
    form_class = CustomerProductForm
    paginate_by = 50
    queryset = CustomerProduct.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("customerproduct_list")
    search_lookup_fields = ["product__title", "customer__name", "agent__full_name"]


class CustomerProductList(CustomerProductMixin, ListView):
    template_name = "customerproduct/customerproduct_list.html"
    queryset = CustomerProduct.objects.filter(status=True, is_deleted=False)


class CustomerProductDetail(CustomerProductMixin, DetailView):
    template_name = "customerproduct/customerproduct_detail.html"


class CustomerProductCreate(CustomerProductMixin, CreateView):
    template_name = "create.html"

    def form_valid(self, form):
        form.instance.agent = self.request.user
        return super().form_valid(form)


class CustomerProductUpdate(CustomerProductMixin, UpdateView):
    template_name = "update.html"

    def form_valid(self, form):
        form.instance.agent = self.request.user
        return super().form_valid(form)


class CustomerProductDelete(CustomerProductMixin, DeleteMixin, View):
    pass

'''  STock VIews '''

from .models import ProductStock
from .forms import ProductStockForm

class ProductStockMixin:
    model = ProductStock
    form_class = ProductStockForm
    paginate_by = 20
    queryset = ProductStock.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('productstock_list')


class ProductStockList(ProductStockMixin, ListView):
    template_name = "productstock/productstock_list.html"
    queryset = ProductStock.objects.filter(status=True,is_deleted=False)

class ProductStockDetail(ProductStockMixin, DetailView):
    template_name = "productstock/productstock_detail.html"

class ProductStockCreate(ProductStockMixin, CreateView):
    template_name = "create.html"

class ProductStockUpdate(ProductStockMixin, UpdateView):
    template_name = "update.html"

class ProductStockDelete(ProductStockMixin, DeleteMixin, View):
    pass

"""  ----------------   """
from .models import BranchStock, BranchStockTracking, ItemReconcilationApiItem
from .forms import BranchStockForm
class BranchStockMixin(IsAdminAccountingOrStoreKeeperMixin): 
    model = BranchStock
    form_class = BranchStockForm
    paginate_by = 10
    queryset = BranchStock.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('branchstock_list')

class BranchStockList(BranchStockMixin, ListView):
    template_name = "branchstock/branchstock_list.html"

class BranchStockDetail(BranchStockMixin, DetailView):
    template_name = "branchstock/branchstock_detail.html"

class BranchStockCreate(BranchStockMixin, CreateView):
    template_name = "branchstock/branchstock_create.html"

class BranchStockUpdate(BranchStockMixin, UpdateView):
    template_name = "update.html"

class BranchStockDelete(BranchStockMixin, DeleteMixin, View):
    pass


from django.db.models import Sum
class ReconcileView(View): 

    def get(self, request):
        opening_exists = BranchStockTracking.objects.count()
        if opening_exists <= 0:
            return render(request, 'item_reconcilation/reconcilation.html',{'show_opening':True})

        branch = request.GET.get('branch', None)
        filter_date = request.GET.get('date')
        branches = Branch.objects.all()
        if not branch or not filter_date:
            return render(request, 'item_reconcilation/reconcilation.html',{'message':'Please Select a Branch and Date', 'branches':branches})
        try:
            filter_date = datetime.strptime(filter_date, '%Y-%m-%d').date().strftime('%Y-%m-%d')
        except Exception:
            return render(request, 'item_reconcilation/reconcilation.html',{'message':'Date format must me YYYY-mm-dd', 'branches':branches})

        filter_branch = get_object_or_404(Branch, branch_code__iexact=branch)

        exists_in_bst = BranchStockTracking.objects.filter(date=filter_date, branch=filter_branch).exists()
        if not exists_in_bst:
            products = Product.objects.filter(reconcile=True).order_by('title').values()
            api_items = ItemReconcilationApiItem.objects.filter(date=filter_date, branch=filter_branch, product__reconcile=True).values()
            received = BranchStock.objects.filter(created_at__contains=filter_date, branch=filter_branch, product__reconcile=True, status=True, is_deleted=False).values('product').annotate(quantity=Sum('quantity'))
            bills = Bill.objects.filter(transaction_date=filter_date, branch=filter_branch, status=True)

            new_products = {}
            for product in products:
                for k, v in product.items():
                    if k =='id':
                        physical_count = 0
                        if BranchStockTracking.objects.filter(branch=filter_branch, product_id=v).exists():
                            physical_count = BranchStockTracking.objects.filter(branch=filter_branch, product_id=v).last().physical
                        new_products[str(v)] = {'title':product.get('title'), 'opening': physical_count}
                        break
        
            for item in api_items:
                product_id = str(item.get('product_id'))
                new_products[product_id]['wastage'] = item.get('wastage', 0)
                new_products[product_id]['returned'] = item.get('returned', 0)
                new_products[product_id]['physical'] = item.get('physical', 0)

            for rec in received:
                product_id = str(rec.get('product'))
                new_products[product_id]['received'] = rec.get('quantity')
            
            for bill in bills:
                for item in bill.bill_items.all():
                    product_id = str(item.product.id)
                    if item.product.reconcile:
                        has_sold = new_products[product_id].get('sold', None)
                        if has_sold:
                            new_products[product_id]['sold'] += item.product_quantity
                        else:
                            new_products[product_id]['sold'] = item.product_quantity

            product_to_view = []
            for k,v in new_products.items():
                new_dict = {'id': k, **v}
                if not 'opening' in new_dict:
                    new_dict['opening'] = 0
                if not 'received' in new_dict:
                    new_dict['received'] = 0
                if not 'wastage' in new_dict:
                    new_dict['wastage'] = 0
                if not 'returned' in new_dict:
                    new_dict['returned'] = 0
                if not 'sold' in new_dict:
                    new_dict['sold'] = 0
                if not 'closing' in new_dict:
                    new_dict['closing'] = 0
                if not 'physical' in new_dict:
                    new_dict['physical'] = 0
                if not 'discrepancy' in new_dict:
                    new_dict['discrepancy'] = 0

                product_to_view.append(new_dict)
            
            for prd in product_to_view:
                opening_received = prd.get('opening') + prd.get('received')
                wastage_returned_sold = prd.get('wastage') + prd.get('returned') + prd.get('sold')
                closing_value = opening_received - wastage_returned_sold
                prd['closing'] = closing_value
                prd['discrepancy'] = prd.get('physical') - closing_value

            context = {
                'products':product_to_view,
                'branches':branches,
                'should_save':True,
                'opening_exists': opening_exists
            }
            return render(request, 'item_reconcilation/reconcilation.html',context)
        
        # --------------------------

        products = BranchStockTracking.objects.filter(date=filter_date, branch=filter_branch).order_by('product__title')
        context = {
            'products':products,
            'branches':branches,
            'should_save':False,
            'opening_exists': opening_exists
        }
        return render(request, 'item_reconcilation/reconcilation.html', context)
    
    def post(self, request):
        branches = Branch.objects.all()
        branch_code = request.POST.get('branch').lower()
        reconcile_date = request.POST.get('filter_date')
        branch = get_object_or_404(Branch, branch_code__iexact=branch_code)
        today_date = date.today()
        if datetime.strptime(reconcile_date, '%Y-%m-%d').date() > today_date:
            messages.error(request, f"Date must not be greater than {today_date}")
            return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})
        
        if BranchStockTracking.objects.filter(date__gte=reconcile_date, branch=branch).exists():
            messages.error(request, f"Items from date greater than {reconcile_date} exists")
            return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})

        last_bill_in_tracking_date = BranchStockTracking.objects.last().date
        bill_exists = Bill.objects.filter(transaction_date__gt=last_bill_in_tracking_date, transaction_date__lt= reconcile_date ,status=True).exists()
        api_items_exists = ItemReconcilationApiItem.objects.filter(date__gt=last_bill_in_tracking_date,date__lt=reconcile_date).exists()

        if bill_exists or api_items_exists:
            messages.error(request, f"Please reconcile items form previous date/s")
            return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})

        data = request.POST
        for k in data:
            try:
                product_id = int(k)
                details = data.getlist(k)
                BranchStockTracking.objects.create(
                    product_id=product_id,
                    branch=branch,
                    date=reconcile_date,
                    opening=details[0],
                    received=details[1],
                    wastage=details[2],
                    returned=details[3],
                    sold=details[4],
                    closing=details[5],
                    physical=details[6],
                    discrepancy=details[7],
                    )
            except ValueError:
                pass
            except IntegrityError:
                messages.error(request, "Items for Today's date already exists")
                return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})
        return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})


class BranchStockUploadView(View):
    
    def post(self, request):
        if BranchStockTracking.objects.count() > 0:
            messages.error(request, "Opening data already exists!!")
            return redirect(reverse_lazy("reconcile"))
        file = request.FILES.get('file')
        branches = Branch.objects.all()
        branch_dict = {}
        for b in branches:
            branch_dict[b.branch_code.lower()] = b.pk

        wb = load_workbook(file)
        excel_data = list()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                row_data = list()
                for cell in row:
                    if cell.value:
                        row_data.append(str(cell.value))
                row_data.append(sheet.title)
                excel_data.append(row_data)
       
        product_dict = {}
        for d in excel_data:
            if len(d) < 3:
                continue
            if d[0].lower().startswith('date'):
                continue
            try:
                product_title = d[1].lower().strip()
                product = product_dict.get(product_title, None)
                if not product:
                    product_dict[product_title] = Product.objects.get(title__iexact=product_title).pk
                product_id = product_dict.get(product_title)
                branch_id =  branch_dict.get(d[3].lower())
                quantity = int(d[2])
                opening_date = datetime.strptime(d[0][:10], '%Y-%m-%d')
                BranchStockTracking.objects.create(product_id=product_id, branch_id=branch_id, opening=quantity, physical=quantity, date=opening_date)
            except Exception as e:
                print(e)

        return redirect(reverse_lazy("reconcile"))
    

class UpdateDateForReconcilationView(View):

    def post(self, request):
        from_date = request.POST.get('from_date', None)
        to_date = request.POST.get('to_date', None)
        branch = request.POST.get('branch', None)
        if not from_date or not to_date:
            messages.error(request, 'Please Provide both "From date" and "To date"')
            return redirect('/reconcile')
        
        api_item_exists = ItemReconcilationApiItem.objects.filter(date=to_date,branch=branch).exists()
        endday_exists = EndDayRecord.objects.filter(date=to_date, branch=branch).exists()

        if api_item_exists or endday_exists:
            messages.error(request, 'Records in api item or end day record already exists')
            return redirect('/reconcile')
        branch = get_object_or_404(Branch, pk=branch)
        ItemReconcilationApiItem.objects.filter(date=from_date,branch=branch).update(date=to_date)
        EndDayRecord.objects.filter(date=from_date, branch=branch).update(date=to_date)
        messages.success(request, 'Date has been updated')
        return redirect('/reconcile')